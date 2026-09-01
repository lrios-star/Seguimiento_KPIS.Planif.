"""
Carga y preprocesamiento del archivo Excel de producción y mermas.

Responsabilidades:
- Leer el Excel con todas sus hojas.
- Detectar hojas que corresponden a periodos semanales (formato DD-MM).
- Parsear las fechas de periodo y ordenar cronológicamente.
- Extraer la tabla de datos y el resumen lateral de cada hoja.
- Proveer funciones cacheadas para Streamlit.
"""
from __future__ import annotations

import re
from datetime import datetime, date
from typing import Optional

import pandas as pd
import openpyxl

from config import EXCEL_PATH, EXPECTED_COLUMNS, NUMERIC_COLUMNS, SUMMARY_LABELS

# Regex para detectar nombres de hoja con formato "DD-MM" (con posible espacio)
_PERIOD_PATTERN = re.compile(r"^\s*(\d{1,2})-(\d{1,2})\s*$")


def _parse_period_name(name: str, reference_year: Optional[int] = None) -> Optional[date]:
    """Convierte un nombre de hoja tipo 'DD-MM' en una fecha.

    El año se determina de la siguiente forma:
    1. Si se provee reference_year, se usa ese.
    2. Si no, se usa el año actual.

    Gestiona el cruce de año (ej. diciembre → enero).
    """
    match = _PERIOD_PATTERN.match(name)
    if not match:
        return None

    day = int(match.group(1))
    month = int(match.group(2))

    if reference_year is None:
        reference_year = datetime.now().year

    try:
        return date(reference_year, month, day)
    except ValueError:
        return None


def _detect_reference_year(wb: openpyxl.Workbook) -> int:
    """Intenta determinar el año de referencia leyendo la celda INICIO de alguna hoja."""
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row_idx in range(1, min(5, ws.max_row + 1)):
            cell_a = ws.cell(row=row_idx, column=1).value
            cell_b = ws.cell(row=row_idx, column=2).value
            if cell_a and str(cell_a).strip().upper() == "INICIO" and isinstance(cell_b, datetime):
                return cell_b.year
    return datetime.now().year


def discover_periods(excel_path=None) -> list[dict]:
    """Descubre las hojas que representan periodos semanales.

    Returns:
        Lista de dicts ordenada cronológicamente:
        [{"sheet_name": "17-08 ", "period_date": date(2026,8,17), "label": "17-08"}, ...]
    """
    path = excel_path or EXCEL_PATH
    wb = openpyxl.load_workbook(path, read_only=True)
    ref_year = _detect_reference_year(wb)
    wb.close()

    # Reabrir para leer nombres (read_only a veces tiene issues)
    wb = openpyxl.load_workbook(path, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    periods = []
    for name in sheet_names:
        period_date = _parse_period_name(name, reference_year=ref_year)
        if period_date is not None:
            periods.append({
                "sheet_name": name,
                "period_date": period_date,
                "label": name.strip(),
            })

    # Ordenar cronológicamente
    periods.sort(key=lambda p: p["period_date"])
    return periods


def _find_op_column(header_row: pd.Series) -> Optional[int]:
    """Encuentra el índice de la columna 'OP' en la fila de encabezado."""
    for idx, val in enumerate(header_row.values):
        if isinstance(val, str) and val.strip().upper() == "OP":
            return idx
    return None


def load_sheet_data(sheet_name: str, excel_path=None) -> pd.DataFrame:
    """Lee la tabla de datos de una hoja específica.

    Detecta dinámicamente la posición de la columna OP para manejar
    hojas con diferente número de columnas a la izquierda.

    Returns:
        DataFrame con las columnas estandarizadas.
    """
    path = excel_path or EXCEL_PATH
    df_raw = pd.read_excel(path, sheet_name=sheet_name, header=None)

    # Encontrar la columna OP en la fila de encabezado (fila 0)
    header_row = df_raw.iloc[0]
    op_col = _find_op_column(header_row)

    if op_col is None:
        raise ValueError(
            f"No se encontró la columna 'OP' en la hoja '{sheet_name}'. "
            f"Encabezados encontrados: {list(header_row.dropna().values)}"
        )

    # Extraer las columnas de datos desde OP en adelante
    data_cols = list(range(op_col, df_raw.shape[1]))
    df = df_raw.iloc[1:, data_cols].copy()
    col_names = [str(header_row.iloc[c]).strip() if pd.notna(header_row.iloc[c]) else f"col_{c}" for c in data_cols]
    df.columns = col_names

    # Filtrar filas que tienen OP (descartar filas vacías)
    df = df.dropna(subset=["OP"])

    # Convertir OP a int
    df["OP"] = pd.to_numeric(df["OP"], errors="coerce")
    df = df.dropna(subset=["OP"])
    df["OP"] = df["OP"].astype(int)

    # Convertir columnas numéricas
    for col in NUMERIC_COLUMNS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # Convertir fecha
    if "FECHA DE CIERRE" in df.columns:
        df["FECHA DE CIERRE"] = pd.to_datetime(df["FECHA DE CIERRE"], errors="coerce")

    # Reset index
    df = df.reset_index(drop=True)

    return df


def load_sheet_summary(sheet_name: str, excel_path=None) -> dict:
    """Lee el resumen de indicadores del panel izquierdo de la hoja.

    Returns:
        Dict con claves normalizadas:
        {"ops_cerradas": 90, "mt_planificados": 768331, ...}
    """
    path = excel_path or EXCEL_PATH
    df_raw = pd.read_excel(path, sheet_name=sheet_name, header=None)

    summary = {}
    for row_idx in range(min(16, len(df_raw))):
        a_val = df_raw.iloc[row_idx, 0]
        b_val = df_raw.iloc[row_idx, 1]
        if pd.notna(a_val):
            key = str(a_val).strip().upper()
            if key in SUMMARY_LABELS:
                normalized = SUMMARY_LABELS[key]
                try:
                    summary[normalized] = float(b_val) if pd.notna(b_val) else 0
                except (ValueError, TypeError):
                    summary[normalized] = 0

    # Extraer fechas de inicio/fin
    for row_idx in range(min(5, len(df_raw))):
        a_val = df_raw.iloc[row_idx, 0]
        b_val = df_raw.iloc[row_idx, 1]
        if pd.notna(a_val):
            label = str(a_val).strip().upper()
            if label == "INICIO":
                summary["inicio"] = b_val
            elif label == "FIN":
                summary["fin"] = b_val

    return summary


def load_all_periods(excel_path=None) -> dict:
    """Carga datos y resúmenes de todos los periodos disponibles.

    Returns:
        Dict con estructura:
        {
            "periods": [...],  # lista ordenada de periodos
            "data": {"17-08": DataFrame, ...},
            "summaries": {"17-08": dict, ...},
        }
    """
    path = excel_path or EXCEL_PATH
    periods = discover_periods(path)

    if not periods:
        return {"periods": [], "data": {}, "summaries": {}, "details": {}}

    wb = openpyxl.load_workbook(path, read_only=True)
    exact_sheet_names = wb.sheetnames
    wb.close()

    data = {}
    summaries = {}
    details = {}
    errors = {}

    for period in periods:
        label = period["label"]
        sheet = period["sheet_name"]
        try:
            data[label] = load_sheet_data(sheet, path)
            summaries[label] = load_sheet_summary(sheet, path)
            
            # Buscar hoja de detalle correspondiente
            target_detail_name = f"DETALLE MERMAS {label}".upper().strip()
            for s_name in exact_sheet_names:
                if s_name.upper().strip() == target_detail_name:
                    try:
                        df_det = pd.read_excel(path, sheet_name=s_name)
                        if "OP" in df_det.columns:
                            # Forward fill solo para OP para mantener la agrupación lógica
                            df_det["OP"] = df_det["OP"].ffill()
                            df_det["OP"] = pd.to_numeric(df_det["OP"], errors="coerce")
                        details[label] = df_det
                    except Exception as e:
                        errors[f"{label}_detalle"] = str(e)
                    break

        except Exception as e:
            errors[label] = str(e)

    return {
        "periods": periods,
        "data": data,
        "summaries": summaries,
        "details": details,
        "errors": errors,
    }
